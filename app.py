from flask import Flask, request, render_template
from collections import defaultdict
import csv
from openpyxl import load_workbook
import io

app = Flask(__name__)

# -------------------------------
# Helper Functions
# -------------------------------

def safe_int(value):
    try:
        return int(float(value))
    except:
        return 0

def safe_float(value):
    try:
        return float(value)
    except:
        return 0

def parse_csv(file):
    """Parse the uploaded CSV or XLSX into 4 sections"""
    my_hitters = []
    my_pitchers = []
    opp_hitters = []
    opp_pitchers = []

    try:
        if file.filename.endswith(".xlsx"):
            wb = load_workbook(filename=io.BytesIO(file.read()), data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                section = sheet_name.upper()
                data = [dict(zip(headers, row)) for row in rows]
                
                if "MY_HITTER" in section:
                    my_hitters = data
                elif "MY_PITCHER" in section:
                    my_pitchers = data
                elif "OPP_HITTER" in section:
                    opp_hitters = data
                elif "OPP_PITCHER" in section:
                    opp_pitchers = data

        else:
            # CSV file
            file.stream.seek(0)
            reader = csv.DictReader(io.StringIO(file.read().decode()))
            if "SECTION" not in reader.fieldnames:
                raise ValueError("CSV must have a 'SECTION' column.")
            for row in reader:
                section = str(row["SECTION"]).strip().upper()
                if section == "MY HITTERS":
                    my_hitters.append(row)
                elif section == "MY PITCHERS":
                    my_pitchers.append(row)
                elif section == "OPP HITTERS":
                    opp_hitters.append(row)
                elif section == "OPP PITCHERS":
                    opp_pitchers.append(row)

    except Exception as e:
        raise ValueError(f"Error reading file: {str(e)}")

    return my_hitters, my_pitchers, opp_hitters, opp_pitchers

def score_player(player, opp_pitcher=None):
    """Simple hitter scoring formula"""
    avg = safe_float(player.get("Avg", 0))
    hr = safe_int(player.get("HR", 0))
    rbi = safe_int(player.get("RBI", 0))
    obp = safe_float(player.get("OBP", 0))
    score = avg * 100 + hr * 5 + rbi * 2 + obp * 100
    return score

def build_lineup(my_hitters, opp_pitchers):
    opp_pitcher = opp_pitchers[0] if opp_pitchers else {}
    by_pos = defaultdict(list)
    for p in my_hitters:
        if "POS" in p:
            by_pos[p["POS"]].append(p)

    lineup = []
    for pos, players in by_pos.items():
        best_player = max(players, key=lambda x: score_player(x, opp_pitcher))
        best_player["Score"] = round(score_player(best_player, opp_pitcher), 1)
        lineup.append(best_player)
    return lineup

# -------------------------------
# Flask Route
# -------------------------------

@app.route("/", methods=["GET", "POST"])
def index():
    lineup = None
    rotation = None
    error = None

    if request.method == "POST":
        try:
            if "stats_file" not in request.files:
                raise ValueError("No file uploaded.")

            file = request.files["stats_file"]
            if file.filename == "":
                raise ValueError("No file selected.")

            my_hitters, my_pitchers, opp_hitters, opp_pitchers = parse_csv(file)

            if not my_hitters:
                raise ValueError("No valid hitter stats found for your team.")
            if not my_pitchers:
                raise ValueError("No valid pitcher stats found for your team.")
            if not opp_pitchers:
                raise ValueError("No valid opponent pitcher stats found.")

            lineup = build_lineup(my_hitters, opp_pitchers)
            rotation = sorted(my_pitchers, key=lambda p: safe_float(p.get("ERA", 99)))[:3]

        except Exception as e:
            error = f"Error processing file: {str(e)}"

    return render_template("index.html", lineup=lineup, rotation=rotation, error=error)

# -------------------------------
# Run Flask
# -------------------------------

import os
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
